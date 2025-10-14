from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.contrib import messages
from django.views.generic import ListView
from django.contrib.auth.hashers import make_password
from django.contrib.admin.models import LogEntry, CHANGE
from django.contrib.contenttypes.models import ContentType
from .models import Grado, Periodo, Paralelo, PadronElectoral, CredencialUsuario
from Aplicaciones.elecciones.models import Lista, Candidato
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import io
import random
import string
from django.core.mail import send_mail
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin



#CRUD GRADOS
@login_required(login_url='agregarLogin')
def generar_credenciales(request):
    # Inicializar el diccionario de credenciales generadas
    if not hasattr(request, 'session'):
        request.session = {}
    if 'credenciales_generadas' not in request.session:
        request.session['credenciales_generadas'] = {}
    
    if request.method == 'POST':
        # Obtener todos los usuarios del padrón que no tienen credencial
        usuarios_sin_credencial = PadronElectoral.objects.filter(
            credencial__isnull=True
        )
        
        credenciales_generadas = []
        
        for padron in usuarios_sin_credencial:
            # Generar una contraseña aleatoria de 8 caracteres
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            
            try:
                # Generar una contraseña aleatoria de 8 caracteres
                password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                
                # Crear la nueva credencial con la contraseña en texto plano
                credencial = CredencialUsuario(
                    padron=padron,
                    usuario=padron.cedula,
                    estado='activo',
                    _contrasena_plana=password  # Establecer la contraseña en texto plano
                )
                
                # Guardar el objeto para generar el hash
                credencial.save()  # Esto generará automáticamente el hash en contrasena_encriptada
                
                # Guardar la contraseña en la sesión
                request.session['credenciales_generadas'][str(credencial.id)] = password
                request.session.modified = True
                
                credenciales_generadas.append({
                    'id': credencial.id,
                    'nombre': f"{padron.nombre} {padron.apellidos}",
                    'cedula': padron.cedula,
                    'contrasena': password,
                    'correo': padron.correo,
                    'estado': 'activo'
                })
                
                print(f"[DEBUG] Credencial creada para {padron.cedula} con contraseña: {password}")
                
            except Exception as e:
                print(f"[ERROR] Error al crear credencial para {padron.cedula}: {e}")
        
        if credenciales_generadas:
            messages.success(request, f'Se generaron {len(credenciales_generadas)} nuevas credenciales.')
            return render(request, 'padron/credenciales.html', {
                'credenciales': credenciales_generadas,
                'mostrar_credenciales': True
            })
        else:
            messages.info(request, 'No hay usuarios sin credenciales para generar.')
    
    # Si es GET o no se generaron credenciales, mostrar todas las credenciales
    credenciales = []
    for cred in CredencialUsuario.objects.select_related('padron').all():
        # Obtener la contraseña de la sesión si existe
        contrasena_plana = request.session.get('credenciales_generadas', {}).get(str(cred.id), "")
        
        print(f"[DEBUG] Mostrando credencial para {cred.padron.cedula}. Contraseña: {contrasena_plana}")
            
        credenciales.append({
            'id': cred.id,
            'nombre': f"{cred.padron.nombre} {cred.padron.apellidos}",
            'cedula': cred.padron.cedula,
            'contrasena': contrasena_plana,
            'correo': cred.padron.correo,
            'estado': cred.estado
        })
    
    # Mover el return fuera del bucle for
    return render(request, 'padron/credenciales.html', {
        'credenciales': credenciales,
        'mostrar_envio': True
    })

    # Si es GET, mostrar la lista de credenciales existentes
    credenciales = []
    for cred in CredencialUsuario.objects.select_related('padron').all():
        # Asegurarse de que tenemos la contraseña en texto plano
        contrasena_plana = cred.get_contrasena_plana  # Sin paréntesis, es una propiedad
        credenciales.append({
            'id': cred.id,
            'nombre': f"{cred.padron.nombre} {cred.padron.apellidos}",
            'cedula': cred.padron.cedula,
            'contrasena': contrasena_plana,  # Usar la contraseña en texto plano
            'correo': cred.padron.correo,
            'estado': cred.estado
        })
        
    return render(request, 'padron/credenciales.html', {
        'credenciales': credenciales,
        'mostrar_envio': False
    })

@login_required(login_url='agregarLogin')
def exportar_credenciales_pdf(request):
    """
    Genera un PDF con las credenciales de los usuarios
    """
    # Obtener todas las credenciales
    credenciales = CredencialUsuario.objects.select_related('padron').all()
    
    # Crear un objeto BytesIO para el PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="credenciales.pdf"'
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=72, bottomMargin=18)
    
    # Estilos
    styles = getSampleStyleSheet()
    style_heading = styles['Heading1']
    style_body = styles['BodyText']
    style_body.alignment = 1  # Centrado
    
    # Contenido del documento
    elements = []
    
    # Título del documento
    elements.append(Paragraph("Credenciales de Acceso", style_heading))
    elements.append(Spacer(1, 20))
    
    # Crear datos para la tabla
    data = [['N°', 'Cédula', 'Nombre Completo', 'Grado', 'Paralelo', 'Usuario', 'Contraseña']]
    
    for i, cred in enumerate(credenciales, 1):
        contrasena_plana = request.session.get('credenciales_generadas', {}).get(str(cred.id), "")
        grado = cred.padron.grado.nombre if cred.padron.grado else ""
        paralelo = cred.padron.paralelo.nombre if cred.padron.paralelo else ""
        
        data.append([
            str(i),
            cred.padron.cedula,
            f"{cred.padron.nombre} {cred.padron.apellidos}",
            grado,
            paralelo,
            cred.usuario,
            contrasena_plana or "*******"
        ])
    
    # Crear la tabla
    table = Table(data)
    
    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    
    # Aplicar estilos alternados a las filas
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)
    
    table.setStyle(style)
    elements.append(table)
    
    # Construir el PDF
    doc.build(elements)
    
    return response

@login_required(login_url='agregarLogin')
def enviar_credenciales(request):
    if request.method == 'POST':
        print('=== INICIO DE ENVÍO DE CREDENCIALES ===')
        print('POST data:', request.POST)
        print('IDs seleccionados:', request.POST.getlist('credenciales'))
        
        try:
            # Verificar que hay credenciales seleccionadas
            ids_seleccionados = request.POST.getlist('credenciales')
            if not ids_seleccionados:
                messages.error(request, 'Debe seleccionar al menos una credencial para enviar')
                print('=== FIN - NO SE SELECCIONARON CREDENCIALES ===')
                return redirect('generar_credenciales')
            
            try:
                # Obtener las credenciales seleccionadas
                credenciales = CredencialUsuario.objects.filter(id__in=ids_seleccionados).select_related('padron')
                print(f'Número de credenciales encontradas: {credenciales.count()}')
                
                if not credenciales.exists():
                    raise ValueError('No se encontraron credenciales válidas para enviar')
                    
                # Enviar correos electrónicos
                for credencial in credenciales:
                    print(f'=== PROCESANDO CREDENCIAL: {credencial.padron.cedula} ===')
                    print(f'Nombre: {credencial.padron.nombre} {credencial.padron.apellidos}')
                    print(f'Correo: {credencial.padron.correo}')
                    print(f'Estado: {credencial.estado}')
                    
                    try:
                        # Validar el formato del correo
                        from django.core.validators import validate_email
                        try:
                            validate_email(credencial.padron.correo)
                            print('Correo válido')
                        except Exception as e:
                            print(f'Error de validación de correo: {str(e)}')
                            raise ValueError(f'Correo electrónico inválido: {credencial.padron.correo}')
                        
                        # Obtener la contraseña de la sesión
                        contrasena_plana = request.session.get('credenciales_generadas', {}).get(str(credencial.id), "")
                        
                        if not contrasena_plana:
                            print(f"[ADVERTENCIA] No se encontró contraseña en texto plano para la credencial {credencial.id}")
                            contrasena_plana = "[CONTRASEÑA NO DISPONIBLE]"
                        
                        subject = 'Credenciales de acceso al Sistema de Votación de la Unidad Educativa Riobamba'
                        message = f"""Hola {credencial.padron.nombre} {credencial.padron.apellidos},

Tus credenciales de acceso al sistema son las siguientes:

Usuario (Cédula): {credencial.padron.cedula}
Contraseña: {contrasena_plana}

Atentamente,
Consejo Electoral - Unidad Educativa Riobamba"""
                        
                        print('Enviando correo...')
                        send_mail(
                            subject,
                            message,
                            settings.DEFAULT_FROM_EMAIL,  # Usar el remitente configurado en settings.py
                            [credencial.padron.correo],
                            fail_silently=False,  # Cambiado a False para ver errores
                        )
                        print('Correo enviado exitosamente')
                        
                        # No cambiar el estado al enviar el correo
                        # El estado se actualizará cuando el usuario cambie su contraseña
                        print('Correo enviado exitosamente sin cambiar el estado')
                        
                    except Exception as e:
                        import traceback
                        print('=== ERROR DETALLADO ===')
                        print(f'Tipo de error: {type(e).__name__}')
                        print(f'Mensaje de error: {str(e)}')
                        print('Traceback:')
                        print(traceback.format_exc())
                        print('=== FIN DEL ERROR ===')
                        
                        messages.warning(request, f'Error al enviar correo a {credencial.padron.correo}: {str(e)}')
                        print(f'Error al enviar correo a {credencial.padron.correo}: {str(e)}')
                        
                # Si no hubo errores críticos, mostrar mensaje de éxito general
                messages.success(request, 'Credenciales enviadas exitosamente. Algunos correos pueden haber fallado.')
                print('=== FIN - PROCESO COMPLETADO ===')
                return redirect('generar_credenciales')
                
            except ValueError as ve:
                print(f'Error de validación: {str(ve)}')
                messages.error(request, str(ve))
                print('=== FIN - ERROR DE VALIDACIÓN ===')
                return redirect('generar_credenciales')
                
        except Exception as e:
            print(f'Error inesperado: {str(e)}')
            import traceback
            print('=== ERROR INESPERADO ===')
            print(f'Tipo de error: {type(e).__name__}')
            print(f'Mensaje de error: {str(e)}')
            print('Traceback:')
            print(traceback.format_exc())
            print('=== FIN DEL ERROR ===')
            messages.error(request, 'Ocurrió un error inesperado al enviar las credenciales')
            print('=== FIN - ERROR INESPERADO ===')
            return redirect('generar_credenciales')
    
    return redirect('generar_credenciales')

#CRUD GRADOS

class GradoListView(LoginRequiredMixin, ListView):
    login_url = 'agregarLogin'
    model = Grado
    template_name = 'grados/agregarGrado.html'
    context_object_name = 'grados'
    paginate_by = 10

    def get_queryset(self):
        return Grado.objects.select_related('periodo').prefetch_related('paralelos').order_by('nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # SOLUCIÓN: Usar el último período por fecha de inicio
        context['periodo_actual'] = Periodo.objects.order_by('-fecha_inicio').first()
        return context

@login_required(login_url='agregarLogin')
def agregar_grado(request):
    # SOLUCIÓN: Obtener el período más reciente por fecha de inicio
    periodo_actual = Periodo.objects.order_by('-fecha_inicio').first()
    
    if not periodo_actual:
        messages.error(request, 'No hay períodos configurados en el sistema')
        return redirect('listar_grados')
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        
        if not nombre:
            messages.error(request, 'El nombre del grado es obligatorio')
            return render(request, 'grados/agregarGrado.html', {
                'periodo_actual': periodo_actual,
                'nombre': nombre
            })
        
        try:
            Grado.objects.create(
                nombre=nombre,
                periodo=periodo_actual
            )
            messages.success(request, 'Grado creado exitosamente!')
            return redirect('listar_grados')
        except Exception as e:
            messages.error(request, f'Error al crear grado: {str(e)}')
            return render(request, 'grados/agregar_grado.html', {
                'periodo_actual': periodo_actual,
                'nombre': nombre
            })
    
    return render(request, 'grados/agregar_grado.html', {
        'periodo_actual': periodo_actual
    })


@login_required(login_url='agregarLogin')
def editar_grado(request, id):
    grado = get_object_or_404(Grado, pk=id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        
        if not nombre:
            messages.error(request, 'El nombre del grado es obligatorio')
            return render(request, 'grados/editar_grado.html', {
                'grado': grado
            })
        
        try:
            grado.nombre = nombre
            # No modificamos el periodo, se mantiene el que ya tenía
            grado.save()
            messages.success(request, 'Grado actualizado correctamente!')
            return redirect('listar_grados')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
            return render(request, 'grados/editar_grado.html', {
                'grado': grado
            })
    
    return render(request, 'grados/editar_grado.html', {
        'grado': grado
    })

@login_required(login_url='agregarLogin')
def eliminar_grado(request, id):
    grado = get_object_or_404(Grado, pk=id)
    
    try:
        if grado.padronelectoral_set.exists():
            messages.error(request, 'No se puede eliminar: Existen estudiantes asociados')
        elif grado.paralelos.exists():
            grado.paralelos.all().delete()
            grado.delete()
            messages.success(request, 'Grado y sus paralelos eliminados correctamente')
        else:
            grado.delete()
            messages.success(request, 'Grado eliminado correctamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar: {str(e)}')
    
    return redirect('listar_grados')

#CRUD PARALELOS

class ParaleloListView(LoginRequiredMixin, ListView):
    login_url = 'agregarLogin'
    model = Paralelo
    template_name = 'paralelos/agregarParalelo.html'
    context_object_name = 'paralelos'
    paginate_by = 15

    def get_queryset(self):
        return Paralelo.objects.select_related('grado', 'grado__periodo').order_by('grado__nombre', 'nombre')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grados'] = Grado.objects.all()
        return context

@login_required(login_url='agregarLogin')
def agregar_paralelo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre').upper()
        grado_id = request.POST.get('grado')
        
        if not nombre or not grado_id:
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('listar_paralelos')
        
        try:
            Paralelo.objects.create(
                nombre=nombre,
                grado_id=grado_id
            )
            messages.success(request, 'Paralelo creado exitosamente!')
        except IntegrityError:
            messages.error(request, 'Este paralelo ya existe para el grado seleccionado')
        except Exception as e:
            messages.error(request, f'Error al crear paralelo: {str(e)}')
        
        return redirect('listar_paralelos')
    
    return render(request, 'paralelos/agregarParalelo.html', {
        'grados': Grado.objects.all()
    })

@login_required(login_url='agregarLogin')
def editar_paralelo(request, id):
    paralelo = get_object_or_404(Paralelo, pk=id)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre').upper()
        grado_id = request.POST.get('grado')
        
        if not nombre or not grado_id:
            messages.error(request, 'Todos los campos son obligatorios')
            return redirect('listar_paralelos')
        
        try:
            paralelo.nombre = nombre
            paralelo.grado_id = grado_id
            paralelo.save()
            messages.success(request, 'Paralelo actualizado correctamente!')
        except IntegrityError:
            messages.error(request, 'Este paralelo ya existe para el grado seleccionado')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {str(e)}')
        
        return redirect('listar_paralelos')
    
    return render(request, 'paralelo/modals/form_paralelo.html', {
        'paralelo': paralelo,
        'grados': Grado.objects.all()
    })

@login_required(login_url='agregarLogin')
def eliminar_paralelo(request, id):
    paralelo = get_object_or_404(Paralelo, pk=id)
    
    if paralelo.padronelectoral_set.exists():
        messages.error(request, '¡No se puede eliminar! Existen estudiantes asociados')
    else:
        try:
            paralelo.delete()
            messages.success(request, 'Paralelo eliminado correctamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar: {str(e)}')
    
    return redirect('listar_paralelos')



# CRUD PADRON ELECTORAL
ESTADOS = [
    ('activo', 'Activo'),
    ('inactivo', 'Inactivo'),
]
@login_required
@login_required(login_url='agregarLogin')
def gestion_padron(request):
    padron = PadronElectoral.objects.select_related('grado', 'paralelo', 'periodo').all().order_by('apellidos', 'nombre')
    grados = Grado.objects.all().prefetch_related('paralelos')
    periodos = Periodo.objects.all()
    paralelos = Paralelo.objects.all()
    
    # Obtener el período actual (el más reciente por fecha de inicio)
    periodo_actual = Periodo.objects.order_by('-fecha_inicio').first()
    
    context = {
        'padron': padron,
        'grados': grados,
        'periodos': periodos,
        'paralelos': paralelos,
        'ESTADOS': ESTADOS,
        'periodo_actual': periodo_actual,  # Añadimos el período actual al contexto
    }
    return render(request, 'padron/agregarPadron.html', context)

@login_required(login_url='agregarLogin')
def agregar_estudiante(request):
    if request.method == 'POST':
        # Obtener datos del formulario
        cedula = request.POST.get('cedula', '').strip()
        nombre = request.POST.get('nombre', '').strip().upper()
        apellidos = request.POST.get('apellidos', '').strip().upper()
        correo = request.POST.get('correo', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        grado_id = request.POST.get('grado')
        paralelo_id = request.POST.get('paralelo')
        periodo_id = request.POST.get('periodo_id')
        estado = request.POST.get('estado')
        
        # Validación básica
        campos_requeridos = {
            'Cédula': cedula,
            'Nombres': nombre,
            'Apellidos': apellidos,
            'Correo electrónico': correo,
            'Grado': grado_id,
            'Paralelo': paralelo_id,
            'Estado': estado
        }
        
        # Verificar campos vacíos
        campos_vacios = [campo for campo, valor in campos_requeridos.items() if not valor]
        if campos_vacios:
            messages.error(request, f'Los siguientes campos son obligatorios: {", ".join(campos_vacios)}')
            return redirect('gestion_padron')
        
        try:
            with transaction.atomic():
                # Verificar si el estudiante ya existe en el mismo período
                periodo = Periodo.objects.get(id=periodo_id) if periodo_id else Periodo.objects.order_by('-fecha_inicio').first()
                if not periodo:
                    raise ValueError('No hay períodos definidos en el sistema')
                
                # Verificar si ya existe un estudiante con la misma cédula en cualquier período
                if PadronElectoral.objects.filter(cedula=cedula).exists():
                    messages.error(request, f'Ya existe un estudiante con la cédula {cedula} en el sistema')
                    return redirect('gestion_padron')
                
                # Verificar si el correo ya está en uso
                if PadronElectoral.objects.filter(correo__iexact=correo).exists():
                    messages.error(request, f'El correo electrónico {correo} ya está registrado')
                    return redirect('gestion_padron')
                
                # Verificar que el grado y paralelo existan
                grado = Grado.objects.get(id=grado_id)
                paralelo = Paralelo.objects.get(id=paralelo_id)
                
                # Verificar que el paralelo pertenezca al grado
                if int(paralelo.grado_id) != int(grado_id):
                    messages.error(request, f'El paralelo {paralelo.nombre} no pertenece al grado {grado.nombre}')
                    # Agregar información de depuración
                    print(f"[DEBUG] Error de validación - Grado ID: {grado_id} (tipo: {type(grado_id)}), Paralelo Grado ID: {paralelo.grado_id} (tipo: {type(paralelo.grado_id)})")
                    return redirect('gestion_padron')
                
                # Crear el estudiante
                estudiante = PadronElectoral.objects.create(
                    cedula=cedula,
                    nombre=nombre,
                    apellidos=apellidos,
                    correo=correo,
                    telefono=telefono,
                    grado=grado,
                    paralelo=paralelo,
                    periodo=periodo,
                    estado=estado
                )
                
                messages.success(request, f'Estudiante {nombre} {apellidos} agregado exitosamente al padrón electoral')
                return redirect('gestion_padron')
                
        except Grado.DoesNotExist:
            messages.error(request, 'El grado seleccionado no existe')
        except Paralelo.DoesNotExist:
            messages.error(request, 'El paralelo seleccionado no existe')
        except Periodo.DoesNotExist:
            messages.error(request, 'El período seleccionado no existe')
        except Exception as e:
            messages.error(request, f'Error al agregar estudiante: {str(e)}')
        
        return redirect('gestion_padron')
    
    # Si no es POST, redirigir a la gestión de padrón
    return redirect('gestion_padron')

@login_required(login_url='agregarLogin')
def editar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(PadronElectoral, id=estudiante_id)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cedula = request.POST.get('cedula')
            nombre = request.POST.get('nombre').upper()  # Consistente con agregar_estudiante
            apellidos = request.POST.get('apellidos').upper()
            correo = request.POST.get('correo')
            telefono = request.POST.get('telefono')
            estado = request.POST.get('estado')  # Campo crítico
            grado_id = request.POST.get('grado')
            paralelo_id = request.POST.get('paralelo')
            periodo_id = request.POST.get('periodo_id')  # Ajustado para consistencia

            # Validación básica
            if not all([cedula, nombre, apellidos, correo, grado_id, paralelo_id, estado]):
                messages.error(request, 'Todos los campos obligatorios deben ser completados')
                return redirect('gestion_padron')

            # Validar el campo estado
            if hasattr(PadronElectoral, 'ESTADOS'):  # Verificar si el modelo tiene ESTADOS definidos
                estados_validos = [estado_val[0] for estado_val in PadronElectoral.ESTADOS]
                if estado not in estados_validos:
                    messages.error(request, f'El estado "{estado}" no es válido')
                    return redirect('gestion_padron')
            else:
                # Si no hay ESTADOS definidos, asumir valores por defecto (ajusta según tu modelo)
                estados_validos = ['activo', 'inactivo']
                if estado not in estados_validos:
                    messages.error(request, f'El estado "{estado}" no es válido')
                    return redirect('gestion_padron')

            # Asignar valores al estudiante
            estudiante.cedula = cedula
            estudiante.nombre = nombre
            estudiante.apellidos = apellidos
            estudiante.correo = correo
            estudiante.telefono = telefono
            estudiante.estado = estado  # Asignar el estado validado
            estudiante.grado = Grado.objects.get(id=grado_id)
            estudiante.paralelo = Paralelo.objects.get(id=paralelo_id)

            # Manejo del período (consistente con agregar_estudiante)
            if not periodo_id:
                periodo = Periodo.objects.order_by('-fecha_inicio').first()
                if not periodo:
                    messages.error(request, 'No hay períodos definidos en el sistema')
                    return redirect('gestion_padron')
                estudiante.periodo = periodo
            else:
                estudiante.periodo = Periodo.objects.get(id=periodo_id)

            # Guardar cambios
            estudiante.save()
            messages.success(request, 'Estudiante actualizado exitosamente!')
        
        except Grado.DoesNotExist:
            messages.error(request, 'El grado seleccionado no existe')
        except Paralelo.DoesNotExist:
            messages.error(request, 'El paralelo seleccionado no existe')
        except Periodo.DoesNotExist:
            messages.error(request, 'El período seleccionado no existe')
        except Exception as e:
            messages.error(request, f'Error al actualizar estudiante: {str(e)}')
        
        return redirect('gestion_padron')

    return render(request, 'editar_estudiante.html', {
        'estudiante': estudiante,
        'grados': Grado.objects.all(),
        'paralelos': Paralelo.objects.all(),
        'periodo_actual': Periodo.objects.order_by('-fecha_inicio').first(),
        'ESTADOS': getattr(PadronElectoral, 'ESTADOS', [('activo', 'Activo'), ('inactivo', 'Inactivo')])
    })
    

@login_required(login_url='agregarLogin')
def eliminar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(PadronElectoral, id=estudiante_id)
    
    try:
        estudiante.delete()
        messages.success(request, 'Estudiante eliminado exitosamente!')
    except Exception as e:
        messages.error(request, f'Error al eliminar estudiante: {str(e)}')
    
    return redirect('gestion_padron')

@login_required(login_url='agregarLogin')
def cargar_paralelos(request):
    grado_id = request.GET.get('grado_id')
    paralelos = Paralelo.objects.filter(grado_id=grado_id).order_by('nombre')
    
    data = {
        'paralelos': [{'id': p.id, 'nombre': p.nombre} for p in paralelos]
    }
    return JsonResponse(data)


@login_required(login_url='agregarLogin')
def estadisticas_padron(request):
    """
    Devuelve estadísticas del padrón en formato JSON para ser usadas en AJAX
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'No autenticado'}, status=403)
        
    total_estudiantes = PadronElectoral.objects.count()
    
    return JsonResponse({
        'total_estudiantes': total_estudiantes,
    })

@login_required(login_url='agregarLogin')
def eliminar_todo_el_padron(request):
    """
    Elimina todo el padrón electoral y los grados/paralelos que ya no están en uso.
    
    Maneja tanto peticiones normales como AJAX.
    """
    if not request.user.is_authenticated:
        error_msg = 'Debe iniciar sesión para realizar esta acción.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': error_msg}, status=403)
        messages.error(request, error_msg)
        return redirect('login')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Contar registros antes de eliminar
                total_estudiantes = PadronElectoral.objects.count()
                
                # 2. Eliminar todos los registros del padrón electoral
                registros_eliminados = PadronElectoral.objects.all().delete()
                
                # 3. Eliminar todos los paralelos
                paralelos_eliminados = Paralelo.objects.all().delete()
                if isinstance(paralelos_eliminados, tuple) and len(paralelos_eliminados) > 0:
                    paralelos_eliminados = paralelos_eliminados[0]  # Tomamos solo el número total de objetos eliminados
                else:
                    paralelos_eliminados = 0
                
                # 4. Eliminar todos los grados
                grados_eliminados = Grado.objects.all().delete()
                if isinstance(grados_eliminados, tuple) and len(grados_eliminados) > 0:
                    grados_eliminados = grados_eliminados[0]  # Tomamos solo el número total de objetos eliminados
                else:
                    grados_eliminados = 0
                
                # 5. Eliminar todos los candidatos (pero mantener las listas)
                candidatos_eliminados = Candidato.objects.all().delete()
                # El método delete() devuelve una tupla: (número_objetos_eliminados, {modelo: número_eliminados})
                if isinstance(candidatos_eliminados, tuple) and len(candidatos_eliminados) > 0:
                    candidatos_eliminados = candidatos_eliminados[0]  # Tomamos solo el número total de objetos eliminados
                else:
                    candidatos_eliminados = 0
                
                # 6. Registrar la acción en el log
                LogEntry.objects.log_action(
                    user_id=request.user.id,
                    content_type_id=ContentType.objects.get_for_model(PadronElectoral).pk,
                    object_id=0,
                    object_repr='Todos los registros del padrón electoral',
                    action_flag=CHANGE,
                    change_message=f'Se eliminaron {total_estudiantes} estudiantes del padrón electoral.\n' 
                                 f'Se eliminaron {paralelos_eliminados} paralelos y {grados_eliminados} grados.\n'
                                 'Se eliminaron todos los candidatos pero se conservaron las listas.'
                )
                
                # Mensaje de éxito
                mensaje = [
                    'Se ha vaciado el padrón electoral exitosamente.',
                    f'• Estudiantes eliminados: {total_estudiantes}',
                    f'• Candidatos eliminados: {candidatos_eliminados if isinstance(candidatos_eliminados, int) else "Todos"}',
                    f'• Se han conservado todas las listas',
                    f'• Paralelos eliminados: {paralelos_eliminados}',
                    f'• Grados eliminados: {grados_eliminados}'
                ]
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': '\n'.join(mensaje)
                    })
                    
                messages.success(request, '\n'.join(mensaje))
                return redirect('gestion_padron')
                
        except Exception as e:
            error_msg = f'Error al eliminar el padrón: {str(e)}\n\nNo se realizaron cambios en la base de datos.'
            
            # Registrar el error en el log del servidor
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error al eliminar el padrón: {str(e)}', exc_info=True)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': error_msg
                }, status=500)
                
            messages.error(request, error_msg)
            return redirect('gestion_padron')
    
    # Si no es una petición POST, redirigir a la página de gestión
    return redirect('gestion_padron')


# FORMATO PADORN ELECTORAL
@login_required(login_url='agregarLogin')
def exportar_padron_excel(request):
    # Crear el libro de trabajo y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "PADRON_ELECTORAL_FORMATO"
    
    # Encabezados
    headers = [
        'Cédula', 'Apellidos', 'Nombres', 'Grado', 'Paralelo', 
        'Correo Electrónico', 'Teléfono(Opcional)'
    ]
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws.column_dimensions[col_letter].width = 20
    
    # Datos
    estudiantes = PadronElectoral.objects.select_related('grado', 'paralelo', 'periodo').all()
    
    for row_num, estudiante in enumerate(estudiantes, 2):
        # Asegurarse de que las columnas coincidan con los encabezados
        ws[f'A{row_num}'] = estudiante.cedula
        ws[f'B{row_num}'] = estudiante.apellidos
        ws[f'C{row_num}'] = estudiante.nombre
        ws[f'D{row_num}'] = estudiante.grado.nombre if estudiante.grado else ''
        ws[f'E{row_num}'] = estudiante.paralelo.nombre if estudiante.paralelo else ''
        ws[f'F{row_num}'] = estudiante.correo if estudiante.correo else ''
        ws[f'G{row_num}'] = estudiante.telefono if estudiante.telefono else ''
    
    # Preparar la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=PADRON_ELECTORAL_FORMATO.xlsx'
    
    # Guardar el libro en la respuesta
    with io.BytesIO() as buffer:
        wb.save(buffer)
        response.write(buffer.getvalue())
    
    return response

# CARGAR EL PADRON ELECTORAL DESDE UN ARCHIVO EXCEL
@login_required(login_url='agregarLogin')
def importar_padron_excel(request):
    print("DEBUG: Iniciando importación de archivo Excel")
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if request.method != 'POST':
        error_msg = 'Método no permitido'
        print(f"DEBUG: Error - {error_msg}")
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg}, status=405)
        messages.error(request, error_msg)
        return redirect('gestion_padron')
        
    if not request.FILES.get('archivo_excel'):
        error_msg = 'No se ha proporcionado ningún archivo'
        print(f"DEBUG: Error - {error_msg}")
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('gestion_padron')
    
    archivo = request.FILES['archivo_excel']
    print(f"DEBUG: Archivo recibido: {archivo.name}, Tamaño: {archivo.size} bytes")
    
    # Verificar que el archivo sea un Excel
    if not archivo.name.endswith(('.xlsx', '.xls')):
        error_msg = 'El archivo debe ser de tipo Excel (.xlsx o .xls)'
        print(f"DEBUG: Error - {error_msg}")
        if is_ajax:
            return JsonResponse({'success': False, 'message': error_msg}, status=400)
        messages.error(request, error_msg)
        return redirect('gestion_padron')
    
    try:
        print("DEBUG: Intentando cargar el archivo Excel...")
        # Cargar el archivo Excel
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
        print(f"DEBUG: Archivo cargado. Hojas: {wb.sheetnames}, Hoja activa: {ws.title}")
        
        # Verificar que el archivo no esté vacío
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"DEBUG: Filas: {max_row}, Columnas: {max_col}")
        
        if max_row <= 1:
            print("DEBUG: Error - El archivo está vacío o solo tiene encabezados")
            messages.error(request, 'El archivo está vacío o solo contiene encabezados')
            return redirect('gestion_padron')
            
        # Obtener encabezados y normalizarlos para hacer coincidencias más flexibles
        headers = [str(cell.value).strip().upper() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"DEBUG: Encabezados encontrados: {headers}")
        
        # Mapeo flexible de columnas
        column_mapping = {
            'cedula': next((i for i, h in enumerate(headers) if 'CEDULA' in h), 0),
            'apellidos': next((i for i, h in enumerate(headers) if 'APELLIDO' in h or 'APELIDO' in h), 1),
            'nombres': next((i for i, h in enumerate(headers) if 'NOMBRE' in h and 'COMPLETO' not in h), 2),
            'grado': next((i for i, h in enumerate(headers) if 'GRADO' in h), 3),
            'paralelo': next((i for i, h in enumerate(headers) if 'PARALELO' in h), 4),
            'correo': next((i for i, h in enumerate(headers) if 'CORREO' in h or 'EMAIL' in h), 5),
            'telefono': next((i for i, h in enumerate(headers) if 'TEL' in h or 'CEL' in h), 6)
        }
        
        # Imprimir el mapeo para depuración
        print(f"DEBUG: Mapeo de columnas: {column_mapping}")
        for key, idx in column_mapping.items():
            print(f"  - {key}: Columna {idx} ('{headers[idx] if idx < len(headers) else 'N/A'}')")
        
        # Verificar períodos académicos existentes
        print("DEBUG: Verificando períodos académicos en la base de datos...")
        todos_periodos = Periodo.objects.all().order_by('-fecha_inicio')
        
        if todos_periodos.exists():
            print("DEBUG: Períodos académicos encontrados:")
            for p in todos_periodos:
                print(f"  - {p.nombre} (ID: {p.id}, Estado: {p.estado}), Fechas: {p.fecha_inicio} a {p.fecha_fin}")
        else:
            print("DEBUG: No se encontraron períodos académicos en la base de datos")
        
        # Obtener el período activo actual (búsqueda no sensible a mayúsculas/minúsculas)
        print("\nDEBUG: Buscando período académico activo...")
        periodo_actual = Periodo.objects.filter(estado__iexact='activo').order_by('-fecha_inicio').first()
        
        if not periodo_actual:
            error_msg = 'No hay un período académico activo. Por favor, cree un período activo primero.'
            print(f"DEBUG: {error_msg}")
            if is_ajax:
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect('gestion_padron')
            
        print(f"DEBUG: Período activo encontrado: {periodo_actual.nombre} (ID: {periodo_actual.id})")
        print(f"DEBUG: Fechas - Inicio: {periodo_actual.fecha_inicio}, Fin: {periodo_actual.fecha_fin}")
            
        print(f"DEBUG: Usando período académico: {periodo_actual.nombre}")
        print(f"DEBUG: Fecha de inicio: {periodo_actual.fecha_inicio}, Fecha de fin: {periodo_actual.fecha_fin}")
        
        # Contadores para estadísticas
        registros_procesados = 0
        registros_omitidos = 0
        
        with transaction.atomic():
            # Empezar desde la segunda fila (asumiendo que la primera es el encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Validar que la fila tenga suficientes columnas (mínimo 6 columnas requeridas)
                    if len(row) < 6:  # Reducido de 7 a 6 ya que el teléfono es opcional
                        registros_omitidos += 1
                        print(f"DEBUG: Fila {row_num} omitida - No tiene suficientes columnas")
                        continue
                        
                    # Obtener los valores de cada columna usando el mapeo
                    try:
                        # Obtener los valores usando el mapeo de columnas
                        cedula = str(row[column_mapping['cedula']]).strip() if row[column_mapping['cedula']] else None
                        apellidos = str(row[column_mapping['apellidos']]).strip().upper() if row[column_mapping['apellidos']] else None
                        nombre = str(row[column_mapping['nombres']]).strip().upper() if row[column_mapping['nombres']] else None
                        grado_nombre = str(row[column_mapping['grado']]).strip().upper() if row[column_mapping['grado']] else None
                        paralelo_nombre = str(row[column_mapping['paralelo']]).strip().upper() if row[column_mapping['paralelo']] else None
                        correo = str(row[column_mapping['correo']]).strip().lower() if row[column_mapping['correo']] else None
                        telefono = str(row[column_mapping['telefono']]).strip() if (len(row) > column_mapping['telefono'] and row[column_mapping['telefono']]) else None
                        
                        print(f"DEBUG: Fila {row_num} - Datos leídos: {cedula}, {apellidos}, {nombre}, {grado_nombre}, {paralelo_nombre}, {correo}, {telefono}")
                    except Exception as e:
                        print(f"DEBUG: Error al procesar fila {row_num}: {str(e)}")
                        registros_omitidos += 1
                        continue
                    
                    # Validar campos obligatorios
                    campos_requeridos = [
                        ('Cédula', cedula),
                        ('Apellidos', apellidos),
                        ('Nombres', nombre),
                        ('Grado', grado_nombre),
                        ('Paralelo', paralelo_nombre),
                        ('Correo', correo)
                    ]
                    
                    campos_faltantes = [nombre for nombre, valor in campos_requeridos if not valor]
                    if campos_faltantes:
                        registros_omitidos += 1
                        print(f"DEBUG: Fila {row_num} omitida - Faltan campos obligatorios: {', '.join(campos_faltantes)}")
                        continue
                    
                    # Validar formato de cédula (permite letras y números, mínimo 4 caracteres)
                    if len(cedula) < 4:
                        print(f"DEBUG: Fila {row_num} omitida - Cédula muy corta (mínimo 4 caracteres): {cedula}")
                        registros_omitidos += 1
                        continue
                        
                    # Validar que solo contenga caracteres alfanuméricos
                    if not cedula.isalnum():
                        print(f"DEBUG: Fila {row_num} omitida - Cédula contiene caracteres no permitidos (solo letras y números): {cedula}")
                        registros_omitidos += 1
                        continue
                    
                    # Validar formato de correo electrónico
                    if '@' not in correo or '.' not in correo.split('@')[-1]:
                        print(f"DEBUG: Fila {row_num} omitida - Correo inválido: {correo}")
                        registros_omitidos += 1
                        continue
                    
                    # Obtener o crear grado y paralelo
                    try:
                        # Limpiar y estandarizar nombres
                        grado_nombre = str(grado_nombre).strip().upper()
                        paralelo_nombre = str(paralelo_nombre).strip().upper()
                        
                        print(f"DEBUG: Procesando - Grado: {grado_nombre}, Paralelo: {paralelo_nombre}")
                        
                        # Obtener o crear el grado
                        grado, created = Grado.objects.get_or_create(
                            nombre=grado_nombre,
                            defaults={'periodo': periodo_actual}
                        )
                        
                        if created:
                            print(f"DEBUG: Nuevo grado creado: {grado.nombre}")
                        
                        # Obtener o crear el paralelo
                        paralelo, created = Paralelo.objects.get_or_create(
                            nombre=paralelo_nombre,
                            grado=grado
                        )
                        
                        if created:
                            print(f"DEBUG: Nuevo paralelo creado: {paralelo.nombre} para el grado {grado.nombre}")
                        
                        # Crear o actualizar estudiante
                        estudiante, created = PadronElectoral.objects.update_or_create(
                            cedula=cedula,
                            defaults={
                                'nombre': nombre,
                                'apellidos': apellidos,
                                'grado': grado,
                                'paralelo': paralelo,
                                'periodo': periodo_actual,
                                'correo': correo,
                                'telefono': telefono,
                                'estado': 'activo'
                            }
                        )
                        
                        if created:
                            print(f"DEBUG: Nuevo estudiante creado: {estudiante.apellidos} {estudiante.nombre}")
                        else:
                            print(f"DEBUG: Estudiante actualizado: {estudiante.apellidos} {estudiante.nombre}")
                        
                        registros_procesados += 1
                        
                    except Exception as e:
                        registros_omitidos += 1
                        continue
                        
                except Exception as e:
                    registros_omitidos += 1
                    continue
            
            # Mensaje de éxito con estadísticas
            if registros_procesados > 0:
                mensaje_exito = {
                    'title': '¡Importación exitosa!',
                    'html': (
                        f'<strong>¡Importación completada con éxito!</strong><br>'
                        f'<strong>Período académico:</strong> {periodo_actual.nombre}<br>'
                        f'<strong>Registros procesados:</strong> {registros_procesados}<br>'
                        f'<strong>Registros omitidos:</strong> {registros_omitidos}'
                    ),
                    'icon': 'success',
                    'registros_procesados': registros_procesados,
                    'registros_omitidos': registros_omitidos
                }
                print(f"DEBUG: {mensaje_exito['html']}")
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': mensaje_exito
                    })
                messages.success(request, mensaje_exito['html'], extra_tags='alert-success')
            else:
                print("DEBUG: No se procesó ningún registro. Verifica los mensajes de error anteriores.")
                mensaje_error = {
                    'title': 'Error en la importación',
                    'html': (
                        '<strong>No se pudo procesar ningún registro. Por favor verifique:</strong><br>'
                        '1. El formato del archivo debe ser: Cédula | Apellidos | Nombres | Grado | Paralelo | Correo | Teléfono (opcional)<br>'
                        '2. La primera fila debe contener los encabezados<br>'
                        '3. No debe haber filas vacías<br>'
                        '4. Las cédulas no deben estar duplicadas<br>'
                        '5. Los correos electrónicos deben tener un formato válido'
                    ),
                    'icon': 'error'
                }
                
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': mensaje_error
                    }, status=400)
                messages.error(request, mensaje_error['html'], extra_tags='alert-danger')
                
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        error_msg = f'Error al procesar el archivo: {str(e)}'
        print(f"DEBUG: {error_msg}")
        print(f"DEBUG: Traceback completo: {error_trace}")
        
        error_response = {
            'title': 'Error en la importación',
            'html': (
                f'<strong>Error al procesar el archivo:</strong> {str(e)}<br><br>'
                '<strong>Por favor verifique que:</strong><br>'
                '1. El archivo no esté abierto en otro programa<br>'
                '2. El formato del archivo sea correcto<br>'
                '3. Tenga los permisos necesarios para leer el archivo'
            ),
            'icon': 'error'
        }
        
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': error_response
            }, status=500)
            
        messages.error(request, error_response['html'])
    
    print("DEBUG: Redirigiendo a la página de gestión de padrón")  # Debug
    return redirect('gestion_padron')
